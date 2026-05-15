import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from src.domain.entities.instrumento_opcional import InstrumentoOpcional, TipoOpcao


TIPO_MAP = {"A": TipoOpcao.AMERICANA, "E": TipoOpcao.EUROPEIA}


def sanitizar_strike(valor_bruto: float, preco_ref: float) -> float:
    """
    Corrige a escala do strike (ex: 134.4 -> 13.44) usando o preço do ativo como referência.
    Prioriza valores que resultem em um strike entre 0.1x e 3.0x o preço da ação.
    """
    if not preco_ref or preco_ref <= 0 or not valor_bruto or valor_bruto <= 0:
        return valor_bruto
    
    import math
    # Testamos uma gama maior de divisores para cobrir layouts variados do Profit
    divisores = [0.01, 0.1, 1.0, 10.0, 100.0, 1000.0]
    melhor_strike = valor_bruto
    menor_score = float('inf')
    
    for d in divisores:
        tentativa = valor_bruto / d
        try:
            # Razão entre o strike e o spot
            razao = tentativa / preco_ref
            
            # Penalizamos fortemente valores fora da faixa 0.1x - 3.0x
            # (Uma opção raramente tem strike 4x o preço da ação ou 0.05x)
            penalidade = 0
            if razao < 0.1 or razao > 3.0:
                penalidade = 10.0 # Grande penalidade
            
            # O score é a distância logarítmica + penalidade
            score = abs(math.log(razao)) + penalidade
            
            if score < menor_score:
                menor_score = score
                melhor_strike = tentativa
        except (ValueError, ZeroDivisionError):
            continue
            
    return melhor_strike


def extrair_strike(codigo: str, preco_ref: float | None = None) -> float | None:
    if not codigo:
        return None
    # Busca a parte numérica no final do código (ex: PETRA300 -> 300)
    nums = re.search(r'(\d+)[a-zA-Z]?$', codigo)
    if not nums:
        return None
    raw = nums.group(1)
    n_len = len(raw)
    
    try:
        val = float(raw)
        
        # Se temos preço de referência, usamos a sanitização inteligente
        if preco_ref and preco_ref > 0:
            return sanitizar_strike(val, preco_ref)
            
        # Fallback para a lógica antiga baseada em tamanho (menos precisa)
        if n_len <= 2:
            return val
        if n_len == 3: # Ex: 300 -> 30.0
            return val / 10.0
        if n_len == 4: # Ex: 3004 -> 30.04
            return val / 100.0
        if n_len >= 5: # Ex: 30040 -> 30.04
            return val / 1000.0
        return val
    except:
        return None


def parse_vencimento(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


class ExcelImporter:
    def __init__(self, filepath: str | Path):
        self.filepath = Path(filepath)

    def importar(self) -> list[InstrumentoOpcional]:
        if not self.filepath.exists():
            raise FileNotFoundError(str(self.filepath))
        wb = openpyxl.load_workbook(str(self.filepath), data_only=True)
        ws = wb.active
        instrumentos = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 5:
                continue
            ativo = row[0]
            vencimento_raw = row[1]
            cod_put = row[2]
            cod_call = row[3]
            tipo_raw = row[4]
            if not ativo or not cod_put or not cod_call:
                continue
            vencimento = parse_vencimento(vencimento_raw)
            if vencimento is None:
                continue
            tipo_str = str(tipo_raw).strip().upper() if tipo_raw else "A"
            tipo_opcao = TIPO_MAP.get(tipo_str, TipoOpcao.AMERICANA)
            instrumentos.append(InstrumentoOpcional(
                ativo=str(ativo).strip(),
                cod_put=str(cod_put).strip(),
                cod_call=str(cod_call).strip(),
                vencimento=vencimento,
                tipo_opcao=tipo_opcao,
            ))
        wb.close()
        return instrumentos
